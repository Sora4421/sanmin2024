# -*- coding: utf-8 -*-
"""
Created on Thu Jun 20 21:02:25 2024

@author: USER
"""

import openpyxl

file = "seles.xlsx"
wb = openpyxl.load_workbook(file,data_only=True)
#ws = wb.active

#指定工作表
ws = wb['2024Q1']
print("目前工作表:",ws.title)

print(ws['A4'].value)
print(ws['C2'].value)
print("總列數:",ws.max_row)
print("總欄數:",ws.max_column)

for i in range(1,ws.max_column+1):
    print(ws.cell(column=i,row=3).value,end=",")
print()




